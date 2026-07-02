import openpyxl
import sys

wb = openpyxl.load_workbook(r'c:\Users\bbqdd\Documents\_my\japaneseproject\Từ-vựng-N5-N1.xlsx')
print('Sheet names:', wb.sheetnames)
sys.stdout.flush()

for s in wb.sheetnames:
    ws = wb[s]
    print(f'Sheet: {s}, Rows: {ws.max_row}, Cols: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            print(list(row))
        else:
            break
    sys.stdout.flush()
