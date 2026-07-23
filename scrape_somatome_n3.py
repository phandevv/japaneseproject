import sys
import urllib.request
import re
import csv
import pandas as pd
from bs4 import BeautifulSoup

sys.stdout.reconfigure(encoding='utf-8')

def extract_clean_text(element):
    if not element:
        return ""
    el = BeautifulSoup(str(element), 'html.parser')
    for tag in el.find_all(['rt', 'rp', 'script', 'style']):
        tag.decompose()
    text = el.get_text(" ", strip=True)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def extract_ruby_text(element):
    if not element:
        return ""
    el = BeautifulSoup(str(element), 'html.parser')
    for script in el.find_all(['script', 'style']):
        script.decompose()
    for ruby in el.find_all('ruby'):
        rt = ruby.find('rt')
        rt_text = rt.get_text(strip=True) if rt else ""
        for tag in ruby.find_all(['rt', 'rp']):
            tag.decompose()
        kanji = ruby.get_text(strip=True)
        if rt_text:
            ruby.replace_with(f"{kanji}({rt_text})")
        else:
            ruby.replace_with(kanji)
    text = el.get_text(" ", strip=True)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def parse_lesson_page(url, week_num, day_num):
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})
    html = urllib.request.urlopen(req).read().decode('utf-8')
    soup = BeautifulSoup(html, 'html.parser')

    h1 = soup.find('h1')
    page_title = extract_clean_text(h1) if h1 else f"第{week_num}週 ({day_num})"

    elements = soup.find_all(['div', 'table'])
    
    items = []
    curr_item = None
    
    for el in elements:
        classes = el.get('class', [])
        
        if 'slide-title' in classes:
            title_text = extract_clean_text(el)
            
            # Reset on practice / answers section
            if any(k in title_text for k in ['練習', 'Đáp án', '実戦問題', 'まとめ', 'Đáp Án']):
                if curr_item and (curr_item['structure'] or curr_item['meaning']):
                    items.append(curr_item)
                curr_item = None
                continue
            
            # Save previous item
            if curr_item and (curr_item['structure'] or curr_item['meaning']):
                items.append(curr_item)
                
            title_with_furigana = extract_ruby_text(el)
            curr_item = {
                'url': url,
                'week': f"Tuần {week_num}",
                'day': f"Ngày {day_num}",
                'page_title': page_title,
                'title': title_text,
                'title_furigana': title_with_furigana,
                'structure': '',
                'meaning': '',
                'explanation': '',
                'examples': []
            }

        elif 'khung' in classes and el.name == 'table':
            if not curr_item:
                curr_item = {
                    'url': url,
                    'week': f"Tuần {week_num}",
                    'day': f"Ngày {day_num}",
                    'page_title': page_title,
                    'title': '',
                    'title_furigana': '',
                    'structure': '',
                    'meaning': '',
                    'explanation': '',
                    'examples': []
                }
            
            rows = el.find_all('tr')
            current_field = None
            for tr in rows:
                row_text = extract_clean_text(tr)
                tds = tr.find_all(['td', 'th'])
                
                if 'Cấu Trúc' in row_text or 'Cấu trúc' in row_text:
                    current_field = 'structure'
                    continue
                elif 'Ý nghĩa' in row_text:
                    current_field = 'meaning'
                    continue
                elif 'Giải thích' in row_text:
                    current_field = 'explanation'
                    continue
                elif 'Ví dụ' in row_text:
                    current_field = 'example_tbl'
                    continue
                
                if current_field:
                    val = extract_clean_text(tr)
                    val = re.sub(r'^[☞\s]+', '', val).strip()
                    if not val:
                        continue
                    if current_field == 'structure' and not curr_item['structure']:
                        curr_item['structure'] = val
                    elif current_field == 'meaning' and not curr_item['meaning']:
                        curr_item['meaning'] = val
                    elif current_field == 'explanation' and not curr_item['explanation']:
                        curr_item['explanation'] = val

        elif 'tudich' in classes and el.name == 'div':
            if curr_item:
                candich = el.find('div', class_='candich')
                nddich = el.find('div', class_='nddich')
                jp_plain = extract_clean_text(candich)
                jp_furi = extract_ruby_text(candich)
                vn = extract_clean_text(nddich)
                if jp_plain or vn:
                    pair = (jp_plain, jp_furi, vn)
                    if pair not in curr_item['examples']:
                        curr_item['examples'].append(pair)
                        
    if curr_item and (curr_item['structure'] or curr_item['meaning']):
        items.append(curr_item)

    return items

def main():
    print("=== BẮT ĐẦU CÀO NGỮ PHÁP N3 SOMATOME TỪ VNJPCLUB ===")
    all_grammar = []
    
    stt = 1
    for w in range(1, 7):
        for d in range(1, 7):
            if w == 1 and d == 1:
                url = "https://www.vnjpclub.com/somatome-n3-ngu-phap/tuan-1.html"
            else:
                url = f"https://www.vnjpclub.com/somatome-n3-ngu-phap/tuan-{w}-{d}.html"
            
            print(f"-> Cào Tuần {w} - Ngày {d} ({url})...")
            try:
                items = parse_lesson_page(url, w, d)
                for it in items:
                    # Format examples as single block text with line breaks
                    ex_blocks = []
                    ex_furi_blocks = []
                    for idx, (jp, jpf, vn) in enumerate(it['examples'], 1):
                        ex_blocks.append(f"{idx}. {jp}\n   👉 {vn}")
                        ex_furi_blocks.append(f"{idx}. {jpf}\n   👉 {vn}")
                    
                    ex_text = "\n".join(ex_blocks)
                    ex_furi_text = "\n".join(ex_furi_blocks)
                    
                    row = {
                        'STT': stt,
                        'Tuần': it['week'],
                        'Ngày': it['day'],
                        'Tên bài học': it['page_title'],
                        'Mẫu ngữ pháp': it['title'],
                        'Mẫu ngữ pháp (Furigana)': it['title_furigana'],
                        'Cấu trúc': it['structure'],
                        'Ý nghĩa': it['meaning'],
                        'Giải thích & Hướng dẫn': it['explanation'],
                        'Ví dụ minh họa': ex_text,
                        'Ví dụ (Furigana)': ex_furi_text,
                        'URL bài học': it['url']
                    }
                    all_grammar.append(row)
                    stt += 1
            except Exception as e:
                print(f"❌ Lỗi khi cào {url}: {e}")

    print(f"\n✅ Đã cào xong tổng cộng {len(all_grammar)} mẫu ngữ pháp N3!")
    
    # Create DataFrame
    df = pd.DataFrame(all_grammar)
    
    # Save to CSV (utf-8-sig for Excel compatibility)
    csv_file = "Ngu_Phap_N3_Somatome.csv"
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"📄 Đã xuất file CSV: {csv_file}")
    
    # Save to XLSX with openpyxl formatting
    xlsx_file = "Ngu_Phap_N3_Somatome.xlsx"
    with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='N3_Grammar')
        
        # Access openpyxl objects to style the sheet
        workbook = writer.book
        worksheet = writer.sheets['N3_Grammar']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Style headers
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Style data cells
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = cell_font
                cell.border = thin_border
                
                # STT, Tuần, Ngày centered
                if col_num in [1, 2, 3]:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
        # Set column widths
        col_widths = {
            'A': 8,   # STT
            'B': 10,  # Tuần
            'C': 10,  # Ngày
            'D': 30,  # Tên bài học
            'E': 25,  # Mẫu ngữ pháp
            'F': 25,  # Mẫu ngữ pháp (Furigana)
            'G': 30,  # Cấu trúc
            'H': 25,  # Ý nghĩa
            'I': 45,  # Giải thích
            'J': 50,  # Ví dụ
            'K': 50,  # Ví dụ Furigana
            'L': 35   # URL
        }
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width

    print(f"📊 Đã xuất file Excel: {xlsx_file}")

if __name__ == '__main__':
    main()
